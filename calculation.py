#CITATION:
#Post, Thierry, Martijn J. van den Assem, Guido Baltussen, and Richard H. Thaler. 2008. "Deal or No Deal? Decision Making under Risk in a Large-Payoff Game Show." American Economic Review, 98 (1): 38-71.

import openpyxl
  
wrkbk = openpyxl.load_workbook("dealorno.xlsx")
sh = wrkbk.active
#turn = 1
#print("ROUND", turn)

list1 = []
list2 = []
list3 = []
list4 = []
list5 = []
list6 = []
list7 = []
list8 = []
list9 = []

for i in range(2, sh.max_row+1):
    
    list = []
    for j in range(12, sh.max_column+1):
        cell_obj = sh.cell(row=i, column=j)
        list.append(cell_obj.value)
    sum = 0
    casesleft = 0
    values = [0.01,1,5,10,25,50,75,100,200,300,400,500,750,1000,5000,10000,25000,50000,75000,100000,200000,300000,400000,500000,750000,1000000]
    for k in range(len(values)):
        if list[k]==1:
            sum += values[k]
            casesleft += 1
    expected = round(sum/casesleft,2)

    offer = sh.cell(row=i, column=11)
    
    percentage = round((offer.value/expected) * 100,2)


    roundnum = sh.cell(row=i, column=9)
    if roundnum.value == 1 and i>2:
        #turn += 1
        #print("ROUND", turn)
        list1.append(percentage)
    
    if roundnum.value == 2:
        list2.append(percentage)

    if roundnum.value == 3:
        list3.append(percentage)

    if roundnum.value == 4:
        list4.append(percentage)

    if roundnum.value == 5:
        list5.append(percentage)

    if roundnum.value == 6:
        list6.append(percentage)

    if roundnum.value == 7:
        list7.append(percentage)

    if roundnum.value == 8:
        list8.append(percentage)

    if roundnum.value == 9:
        list9.append(percentage)

    #print("Row ", i-1, "Offer is ",percentage, "% of expected value")
sum1 = 0
for i in range(len(list1)):
    sum1 += list1[i]
mean1 = sum1/len(list1)
print("Average percent of expected value from round 1: ",mean1)
    
sum2 = 0
for i in range(len(list2)):
    sum2 += list2[i]
mean2 = sum2/len(list2)
print("Average percent of expected value from round 2: ",mean2)    

sum3 = 0
for i in range(len(list3)):
    sum3 += list3[i]
mean3 = sum3/len(list3)
print("Average percent of expected value from round 3: ",mean3)
    
sum4 = 0
for i in range(len(list4)):
    sum4 += list4[i]
mean4 = sum4/len(list4)
print("Average percent of expected value from round 4: ",mean4)

sum5 = 0
for i in range(len(list5)):
    sum5 += list5[i]
mean5 = sum5/len(list5)
print("Average percent of expected value from round 5: ",mean5)
    
sum6 = 0
for i in range(len(list6)):
    sum6 += list6[i]
mean6 = sum6/len(list6)
print("Average percent of expected value from round 6: ",mean6)

sum7 = 0
for i in range(len(list7)):
    sum7 += list7[i]
mean7 = sum7/len(list7)
print("Average percent of expected value from round 7: ",mean7)
    
sum8 = 0
for i in range(len(list8)):
    sum8 += list8[i]
mean8 = sum8/len(list8)
print("Average percent of expected value from round 8: ",mean8)

sum9 = 0
for i in range(len(list9)):
    sum9 += list9[i]
mean9 = sum9/len(list9)
print("Average percent of expected value from round 9: ",mean9)
 